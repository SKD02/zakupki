from io import BytesIO
from typing import Any, Dict, Iterable, List, Optional

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _safe_sheet_title(value: str) -> str:
    title = "".join(ch for ch in str(value or "Шаблон") if ch not in r"[]:*?/\\")[:31]
    return title or "Шаблон"


def build_excel_template_response(
    *,
    filename: str,
    sheet_title: str,
    fields: Iterable[Dict[str, Any]],
    example_row: Optional[Dict[str, Any]] = None,
    description: Optional[str] = None,
) -> StreamingResponse:
    """Создаёт XLSX-шаблон с заголовками, примером и листом описания полей."""
    normalized_fields: List[Dict[str, Any]] = list(fields)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _safe_sheet_title(sheet_title)

    headers = [field.get("label") or field.get("key") for field in normalized_fields]
    worksheet.append(headers)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    required_fill = PatternFill("solid", fgColor="FCE4D6")

    for column_no, field in enumerate(normalized_fields, start=1):
        cell = worksheet.cell(row=1, column=column_no)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.fill = required_fill if field.get("required") else header_fill
        if field.get("required"):
            cell.comment = Comment("Обязательное поле", "Procurement Assistant")

    if example_row:
        worksheet.append([example_row.get(field.get("key")) for field in normalized_fields])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_no, field in enumerate(normalized_fields, start=1):
        values = [str(field.get("label") or field.get("key") or "")]
        if example_row and example_row.get(field.get("key")) is not None:
            values.append(str(example_row.get(field.get("key"))))
        width = min(max(max(len(value) for value in values) + 4, 16), 45)
        worksheet.column_dimensions[get_column_letter(column_no)].width = width

    info = workbook.create_sheet("Описание полей")
    start_row = 1
    if description:
        info.append([description])
        info.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        info.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        info.cell(row=1, column=1).font = Font(bold=True)
        start_row = 3

    for col, title in enumerate(["Поле", "Технический ключ", "Обязательное", "Комментарий"], start=1):
        cell = info.cell(row=start_row, column=col, value=title)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row_no, field in enumerate(normalized_fields, start=start_row + 1):
        info.cell(row=row_no, column=1, value=field.get("label") or field.get("key"))
        info.cell(row=row_no, column=2, value=field.get("key"))
        info.cell(row=row_no, column=3, value="Да" if field.get("required") else "Нет")
        info.cell(row=row_no, column=4, value=field.get("description") or "")

    for column_no, width in enumerate([34, 28, 16, 70], start=1):
        info.column_dimensions[get_column_letter(column_no)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
